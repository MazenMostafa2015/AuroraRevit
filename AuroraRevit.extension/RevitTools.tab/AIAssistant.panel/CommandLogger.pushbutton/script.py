# -*- coding: utf-8 -*-
"""Aurora Command Logger.

Independent pyRevit pushbutton. It does not import or modify AIChat/QuickCommand.
The journal is scanned when this button is opened or refreshed. For true startup
collection, place the same call in a pyRevit startup hook if desired.
"""

from __future__ import print_function

import csv
import datetime
import json
import os
import re
import subprocess

try:
    import clr
    clr.AddReference("PresentationFramework")
    clr.AddReference("PresentationCore")
    clr.AddReference("WindowsBase")
    from System.Windows import Window, Thickness
    from System.Windows.Controls import Button, StackPanel, TextBlock
    from System.Windows.Media import Brushes, SolidColorBrush, Color
    WPF_AVAILABLE = True
except Exception:
    Window = None
    Thickness = None
    Button = None
    StackPanel = None
    TextBlock = None
    Brushes = None
    SolidColorBrush = None
    Color = None
    WPF_AVAILABLE = False

try:
    from pyrevit import forms
except Exception:
    forms = None

DEFAULT_LOG_DIR = r"C:\AuroraRevit_Logs"
CONFIG_PATH = os.path.join(os.environ.get("APPDATA", DEFAULT_LOG_DIR), "AuroraRevit", "command_tools_settings.json")


def _configured_log_dir():
    try:
        with open(CONFIG_PATH, "r") as handle:
            settings = json.load(handle)
            value = settings.get("log_folder") if isinstance(settings, dict) else None
            if value and os.path.isabs(value):
                return os.path.normpath(value)
    except Exception:
        pass
    return DEFAULT_LOG_DIR


LOG_DIR = _configured_log_dir()
XLSX_PATH = os.path.join(LOG_DIR, "CommandLog.xlsx")
CSV_PATH = os.path.join(LOG_DIR, "CommandLog.csv")
STATE_PATH = os.path.join(LOG_DIR, "CommandLog.state.json")
ACCENT = SolidColorBrush(Color.FromRgb(0, 120, 212)) if WPF_AVAILABLE else None
HEADERS = ["Command Name", "Current User", "Timestamp", "Revit Version", "Description/Translation"]

COMMAND_TRANSLATIONS = {
    "ID_EDIT_MOVE": "Move Command",
    "ID_EDIT_COPY": "Copy Command",
    "ID_EDIT_MIRROR": "Mirror Command",
    "ID_EDIT_ROTATE": "Rotate Command",
    "ID_EDIT_ARRAY": "Array Command",
    "ID_EDIT_OFFSET": "Offset Command",
    "ID_EDIT_TRIM": "Trim/Extend Command",
    "ID_EDIT_DELETE": "Delete Command",
    "ID_EDIT_UNDO": "Undo Command",
    "ID_EDIT_REDO": "Redo Command",
    "ID_EDIT_ALIGN": "Align Command",
    "ID_EDIT_SPLIT": "Split Command",
    "ID_EDIT_JOIN": "Join Command",
    "ID_EDIT_PIN": "Pin Command",
    "ID_EDIT_UNPIN": "Unpin Command",
}
COMMAND_RE = re.compile(r"\b(ID_[A-Z0-9_]+)\b")


def _revit_version():
    try:
        return str(__revit__.Application.VersionNumber)
    except Exception:
        return "Unknown"


def _journal_root():
    local_appdata = os.environ.get("LOCALAPPDATA", "")
    if not local_appdata:
        return None
    return os.path.join(local_appdata, "Autodesk", "Revit", "Autodesk Revit " + _revit_version(), "Journals")


def _journal_files():
    root = _journal_root()
    if not root or not os.path.isdir(root):
        return []
    files = []
    for name in os.listdir(root):
        path = os.path.join(root, name)
        if os.path.isfile(path) and name.lower().startswith("journal"):
            try:
                files.append((os.path.getmtime(path), path))
            except Exception:
                pass
    files.sort(reverse=True)
    return [path for _, path in files[:10]]


def _description(command_id):
    return COMMAND_TRANSLATIONS.get(command_id, "Revit command: " + command_id.replace("_", " ").title())


def _read_journal_commands():
    commands = []
    for path in _journal_files():
        try:
            with open(path, "r") as handle:
                text = handle.read()
        except Exception:
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as handle:
                    text = handle.read()
            except Exception:
                continue
        for match in COMMAND_RE.finditer(text):
            command_id = match.group(1)
            # Journal path plus character offset gives each occurrence a stable key.
            commands.append((command_id, path + ":" + str(match.start())))
    return commands


def _ensure_dir():
    if not os.path.isdir(LOG_DIR):
        os.makedirs(LOG_DIR)


def _existing_commands():
    result = set()
    try:
        import openpyxl
        if os.path.isfile(XLSX_PATH):
            workbook = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    result.add(str(row[0]))
            workbook.close()
            return result
    except Exception:
        pass
    if os.path.isfile(CSV_PATH):
        try:
            with open(CSV_PATH, "r") as handle:
                for row in csv.DictReader(handle):
                    if row.get("Command Name"):
                        result.add(row["Command Name"])
        except Exception:
            pass
    return result


def append_log(command_name, description=None, timestamp=None):
    """Append one normalized command entry. Returns the file actually written."""
    _ensure_dir()
    row = [
        command_name,
        os.environ.get("USERNAME", os.environ.get("USER", "Unknown")),
        timestamp or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        _revit_version(),
        description or _description(command_name),
    ]
    try:
        import openpyxl
        workbook = None
        if os.path.isfile(XLSX_PATH):
            workbook = openpyxl.load_workbook(XLSX_PATH)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Command Log"
            sheet.append(HEADERS)
        sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(XLSX_PATH)
        return XLSX_PATH
    except Exception:
        new_file = not os.path.isfile(CSV_PATH)
        with open(CSV_PATH, "a") as handle:
            writer = csv.writer(handle)
            if new_file:
                writer.writerow(HEADERS)
            writer.writerow(row)
        return CSV_PATH


def _load_state():
    try:
        with open(STATE_PATH, "r") as handle:
            value = json.load(handle)
            return set(value if isinstance(value, list) else [])
    except Exception:
        return set()


def _save_state(state):
    try:
        with open(STATE_PATH, "w") as handle:
            json.dump(sorted(list(state)), handle)
    except Exception:
        pass


def refresh_log():
    _ensure_dir()
    processed = _load_state()
    written = 0
    for command_id, occurrence_key in _read_journal_commands():
        if occurrence_key not in processed:
            append_log(command_id)
            processed.add(occurrence_key)
            written += 1
    _save_state(processed)
    return written


def _open_log_folder(_sender=None, _args=None):
    _ensure_dir()
    try:
        os.startfile(LOG_DIR)
    except Exception:
        subprocess.Popen(["explorer.exe", LOG_DIR])


def _show_message(message):
    if forms:
        forms.alert(message, title="Aurora Command Logger")


def show_window():
    count = refresh_log()
    if not WPF_AVAILABLE:
        _show_message("Command logging is active, but the WPF assemblies could not be loaded. The log file is:\n\n" + (XLSX_PATH if os.path.isfile(XLSX_PATH) else CSV_PATH))
        return
    window = Window()
    window.Title = "Aurora Command Logger"
    window.Width = 460
    window.Height = 190
    window.Background = SolidColorBrush(Color.FromRgb(30, 30, 30))
    window.Foreground = Brushes.White
    panel = StackPanel()
    panel.Margin = Thickness(18)
    title = TextBlock(Text="Command log is up to date. Added {0} new journal command(s).".format(count))
    title.Margin = Thickness(0, 0, 0, 16)
    panel.Children.Add(title)
    path = XLSX_PATH if os.path.isfile(XLSX_PATH) else CSV_PATH
    status = TextBlock(Text="File: " + path)
    status.Margin = Thickness(0, 0, 0, 14)
    panel.Children.Add(status)
    buttons = StackPanel()
    buttons.Orientation = 0
    refresh = Button(Content="Refresh Log", Width=120, Height=30, Margin=Thickness(0, 0, 8, 0))
    refresh.Background = ACCENT
    refresh.Click += lambda sender, args: _show_message("Added {0} new command(s).\n\n{1}".format(refresh_log(), path))
    buttons.Children.Add(refresh)
    open_button = Button(Content="Open Log Folder", Width=130, Height=30)
    open_button.Background = ACCENT
    open_button.Click += _open_log_folder
    buttons.Children.Add(open_button)
    panel.Children.Add(buttons)
    window.Content = panel
    window.ShowDialog()


if __name__ == "__main__":
    show_window()
