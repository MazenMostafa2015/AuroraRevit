# -*- coding: utf-8 -*-
"""Aurora Utility Tools.

One independent pyRevit button containing six command utilities. The script is
written for IronPython 2.7 and uses only pyRevit, Revit DB/UI, and optional
openpyxl. Model-changing commands always show a preview and require confirmation.
"""
from __future__ import print_function

import csv
import datetime
import json
import os
import sys
import types

try:
    from pyrevit import forms, revit, DB, UI
except Exception:
    forms = None
    revit = None
    DB = None
    UI = None

DEFAULT_LOG_DIR = r"C:\AuroraRevit_Logs"
STATE_PATH = os.path.join(DEFAULT_LOG_DIR, "AuroraCalculationState.json")
ACCENT = "#FF0078D4"


def _text(value):
    if value is None:
        return ""
    try:
        return str(value)
    except Exception:
        return repr(value)


def _load_logger():
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(os.path.dirname(here), "CommandLogger.pushbutton", "script.py")
    if not os.path.isfile(path):
        return None
    script_dir = os.path.dirname(path)
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)
    module = types.ModuleType("aurora_utility_logger")
    module.__file__ = path
    try:
        with open(path, "rb") as stream:
            source = stream.read()
        exec(compile(source, path, "exec"), module.__dict__)
        return module
    except Exception:
        return None


def _log(command_name, description):
    logger = _load_logger()
    if logger:
        try:
            return logger.append_log(command_name, description)
        except Exception:
            pass
    return None


def _confirm(message, title="Aurora Utility Tools"):
    if not forms:
        return False
    return forms.alert(message, title=title, ok=False, yes=True, no=True)


def _choose(title, options):
    if not forms:
        return None
    try:
        result = forms.CommandSwitchWindow.show(options, message=title)
    except Exception:
        result = forms.SelectFromList.show(options, title=title, button_name="Select")
    if isinstance(result, list):
        return result[0] if result else None
    return result


def _pick_output_folder():
    try:
        return forms.pick_folder(title="Choose output folder")
    except Exception:
        return None


def _ask_filename(default_name):
    try:
        value = forms.ask_for_string(default=default_name, prompt="Output file name", title="Aurora Utility Tools")
        return value.strip() if value else None
    except Exception:
        return default_name


def _printable_view_set(view):
    view_set = DB.ViewSet()
    view_set.Insert(view)
    return view_set


def _select_print_settings(print_manager):
    orientation = _choose("Choose PDF orientation", ["Portrait", "Landscape"])
    paper = _choose("Choose paper size", ["A3", "A4", "Legal"])
    zoom = _choose("Choose zoom", ["Fit to page", "100 percent"])
    if not orientation or not paper or not zoom:
        return None
    return orientation, paper, zoom


def _set_print_parameters(print_manager, orientation, paper_name, zoom):
    """Best-effort settings application across Revit 2023-2025 print APIs."""
    try:
        setup = print_manager.PrintSetup
        setting = setup.InSession
        parameters = setting.PrintParameters
        enum_type = getattr(DB, "PageOrientationType", None)
        if enum_type is not None:
            parameters.PageOrientation = getattr(enum_type, orientation)
        else:
            try:
                parameters.Orientation = orientation
            except Exception:
                pass
        try:
            for paper in print_manager.PaperSizes:
                if paper_name.lower() in _text(paper.Name).lower():
                    parameters.PaperSize = paper
                    break
        except Exception:
            pass
        zoom_type = getattr(DB, "ZoomType", None)
        if zoom == "100 percent" and zoom_type is not None:
            parameters.ZoomType = getattr(zoom_type, "Zoom")
            parameters.Zoom = 100
        elif zoom_type is not None:
            parameters.ZoomType = getattr(zoom_type, "FitToPage")
        try:
            setup.CurrentPrintSetting = setting
        except Exception:
            pass
    except Exception:
        pass


def export_current_view_to_pdf():
    view = revit.doc.ActiveView
    folder = _pick_output_folder()
    if not folder:
        return
    name = _ask_filename("AuroraRevit_{0}.pdf".format(view.Name.replace(" ", "_")))
    if not name:
        return
    if not name.lower().endswith(".pdf"):
        name += ".pdf"
    output_path = os.path.join(folder, name)
    settings = _select_print_settings(revit.doc.PrintManager)
    if not settings:
        return
    orientation, paper, zoom = settings
    preview = "Safe Preview\n\nView: {0}\nOutput: {1}\nOrientation: {2}\nPaper: {3}\nZoom: {4}\n\nSubmit to the configured PDF printer?".format(
        view.Name, output_path, orientation, paper, zoom)
    if not _confirm(preview, "Aurora Export Current View to PDF"):
        return
    try:
        pm = revit.doc.PrintManager
        pm.PrintRange = DB.PrintRange.Select
        _set_print_parameters(pm, orientation, paper, zoom)
        tx = DB.Transaction(revit.doc, "Aurora PDF Print Selection")
        tx.Start()
        try:
            pm.ViewSheetSetting.InSession.Views = _printable_view_set(view)
            tx.Commit()
        except Exception:
            tx.RollBack()
            raise
        try:
            pm.SelectNewPrintDriver("Microsoft Print to PDF")
        except Exception:
            pass
        pm.PrintToFile = True
        pm.PrintToFileName = output_path
        pm.Apply()
        pm.SubmitPrint()
        _log("EXPORT_CURRENT_VIEW_TO_PDF", "Submitted {0} to {1}".format(view.Name, output_path))
        forms.alert("PDF export submitted to the configured Revit printer.", title="Aurora Utility Tools")
    except Exception as error:
        forms.alert("PDF export failed:\n\n" + _text(error), title="Aurora Utility Tools")


def _schedule_matrix(schedule):
    body = schedule.GetTableData().GetSectionData(DB.SectionType.Body)
    columns = body.NumberOfColumns
    rows = body.NumberOfRows
    headers = []
    for column in range(columns):
        heading = "Column {0}".format(column + 1)
        try:
            heading = schedule.Definition.GetField(column).ColumnHeading or heading
        except Exception:
            pass
        headers.append(_text(heading))
    values = []
    for row in range(rows):
        values.append([_text(schedule.GetCellText(DB.SectionType.Body, row, column)) for column in range(columns)])
    return headers, values


def _write_schedule_file(path, headers, values, formatted):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Schedule Export"
        sheet.append(headers)
        for row in values:
            sheet.append(row)
        if formatted:
            fill = PatternFill("solid", fgColor="17365D")
            font = Font(color="FFFFFF", bold=True)
            for cell in sheet[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = 12
            for cell in column:
                width = max(width, min(45, len(_text(cell.value)) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        book.save(path)
        return path
    except Exception:
        csv_path = os.path.splitext(path)[0] + ".csv"
        with open(csv_path, "wb") as stream:
            writer = csv.writer(stream)
            writer.writerow(headers)
            writer.writerows(values)
        return csv_path


def export_active_schedule(formatted):
    schedule = revit.doc.ActiveView
    if not isinstance(schedule, DB.ViewSchedule):
        forms.alert("The active view is not a schedule.", title="Aurora Utility Tools")
        return
    headers, values = _schedule_matrix(schedule)
    folder = _pick_output_folder()
    if not folder:
        return
    default = "{0}_{1}.xlsx".format(schedule.Name.replace(" ", "_"), "Formatted" if formatted else "Export")
    name = _ask_filename(default)
    if not name:
        return
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    path = os.path.join(folder, name)
    preview = "Schedule Preview\n\nSchedule: {0}\nRows: {1}\nColumns: {2}\nOutput: {3}\nFormatting: {4}\n\nExport now?".format(
        schedule.Name, len(values), len(headers), path, "dark blue header" if formatted else "plain")
    if not _confirm(preview, "Aurora Schedule Export"):
        return
    result = _write_schedule_file(path, headers, values, formatted)
    _log("EXPORT_ACTIVE_SCHEDULE_TO_EXCEL", "Exported {0} to {1}".format(schedule.Name, result))
    forms.alert("Export complete:\n\n" + result, title="Aurora Utility Tools")


def _category_options():
    return [
        ("Walls", DB.BuiltInCategory.OST_Walls),
        ("Doors", DB.BuiltInCategory.OST_Doors),
        ("Rooms", DB.BuiltInCategory.OST_Rooms),
        ("Floors", DB.BuiltInCategory.OST_Floors),
        ("Roofs", DB.BuiltInCategory.OST_Roofs),
        ("Windows", DB.BuiltInCategory.OST_Windows),
        ("Mechanical Equipment", DB.BuiltInCategory.OST_MechanicalEquipment),
        ("Ducts", DB.BuiltInCategory.OST_DuctCurves),
        ("Pipes", DB.BuiltInCategory.OST_PipeCurves),
        ("Generic Models", DB.BuiltInCategory.OST_GenericModel),
    ]


def _element_parameters(elements):
    names = set()
    for element in elements[:200]:
        try:
            for parameter in element.Parameters:
                if parameter.Definition and parameter.Definition.Name:
                    names.add(parameter.Definition.Name)
        except Exception:
            pass
    result = list(names)
    result.sort(key=lambda value: value.lower())
    return result


def batch_parameter_translator():
    category_names = [item[0] for item in _category_options()]
    selected_category = _choose("Choose category", category_names)
    if not selected_category:
        return
    category = dict(_category_options())[selected_category]
    elements = list(DB.FilteredElementCollector(revit.doc).OfCategory(category).WhereElementIsNotElementType().ToElements())
    if not elements:
        forms.alert("No instances were found for " + selected_category + ".", title="Aurora Utility Tools")
        return
    parameters = _element_parameters(elements)
    source = _choose("Choose source parameter", parameters)
    target = _choose("Choose target parameter", parameters)
    if not source or not target:
        return
    find_text = forms.ask_for_string(default="W1", prompt="Text to find", title="Batch Parameter Translator")
    replace_text = forms.ask_for_string(default="Wall Type A", prompt="Replacement text", title="Batch Parameter Translator")
    if find_text is None or replace_text is None:
        return
    matches = []
    for element in elements:
        try:
            source_parameter = element.LookupParameter(source)
            value = source_parameter.AsString() if source_parameter else None
            if value and find_text in value:
                matches.append((element, value))
        except Exception:
            pass
    preview = "Safe Preview\n\nCategory: {0}\nSource: {1}\nTarget: {2}\nReplace: {3} -> {4}\nMatching elements: {5}\n\nProceed with a Revit transaction?".format(
        selected_category, source, target, find_text, replace_text, len(matches))
    if not matches or not _confirm(preview, "Aurora Batch Parameter Translator"):
        return
    changed = 0
    skipped = 0
    tx = DB.Transaction(revit.doc, "Aurora Batch Parameter Translator")
    tx.Start()
    try:
        for element, value in matches:
            target_parameter = element.LookupParameter(target)
            if target_parameter and not target_parameter.IsReadOnly:
                target_parameter.Set(value.replace(find_text, replace_text))
                changed += 1
            else:
                skipped += 1
        tx.Commit()
    except Exception:
        tx.RollBack()
        raise
    _log("BATCH_PARAMETER_TRANSLATOR", "Changed {0} {1} elements from {2} to {3}; skipped {4}".format(changed, selected_category, find_text, replace_text, skipped))
    forms.alert("Changed: {0}\nSkipped: {1}".format(changed, skipped), title="Aurora Utility Tools")


def _load_calculation_state():
    try:
        with open(STATE_PATH, "r") as handle:
            return json.load(handle)
    except Exception:
        return None


def _save_calculation_state(state):
    if not os.path.isdir(DEFAULT_LOG_DIR):
        os.makedirs(DEFAULT_LOG_DIR)
    with open(STATE_PATH, "w") as handle:
        json.dump(state, handle, indent=2)


def background_calculation_killer():
    view = revit.doc.ActiveView
    preview = "Safe Performance Preview\n\nActive view: {0}\n\nRevit does not expose a supported global API switch that kills every Color Fill or MEP calculation. Aurora will therefore apply a reversible lightweight display mode to the active view only: Coarse detail and Wireframe where supported. Restore Calculations returns the original settings.\n\nContinue?".format(view.Name)
    if not _confirm(preview, "Aurora Background Calculation Killer"):
        return
    state = {"view_id": view.Id.IntegerValue}
    try:
        state["detail_level"] = _text(view.DetailLevel)
    except Exception:
        pass
    try:
        state["display_style"] = _text(view.DisplayStyle)
    except Exception:
        pass
    tx = DB.Transaction(revit.doc, "Aurora Safe Performance Mode")
    tx.Start()
    try:
        try:
            view.DetailLevel = DB.ViewDetailLevel.Coarse
        except Exception:
            pass
        try:
            view.DisplayStyle = DB.DisplayStyle.Wireframe
        except Exception:
            pass
        tx.Commit()
    except Exception:
        tx.RollBack()
        raise
    _save_calculation_state(state)
    _log("BACKGROUND_CALCULATION_KILLER", "Enabled reversible lightweight display mode for view " + view.Name)
    forms.alert("Safe performance mode enabled for the active view. Use Restore Calculations to revert it.", title="Aurora Utility Tools")


def restore_calculations():
    state = _load_calculation_state()
    if not state:
        forms.alert("No saved performance-mode state was found.", title="Aurora Utility Tools")
        return
    if not _confirm("Restore the saved active-view display settings?", "Aurora Restore Calculations"):
        return
    view = revit.doc.GetElement(DB.ElementId(int(state["view_id"])))
    if not view:
        forms.alert("The original view is no longer available.", title="Aurora Utility Tools")
        return
    tx = DB.Transaction(revit.doc, "Aurora Restore Performance Mode")
    tx.Start()
    try:
        if state.get("detail_level"):
            try:
                view.DetailLevel = getattr(DB.ViewDetailLevel, state["detail_level"].split(".")[-1])
            except Exception:
                pass
        if state.get("display_style"):
            try:
                view.DisplayStyle = getattr(DB.DisplayStyle, state["display_style"].split(".")[-1])
            except Exception:
                pass
        tx.Commit()
    except Exception:
        tx.RollBack()
        raise
    try:
        os.remove(STATE_PATH)
    except Exception:
        pass
    _log("RESTORE_CALCULATIONS", "Restored saved display settings for view " + _text(view.Name))
    forms.alert("Saved display settings restored.", title="Aurora Utility Tools")


def _floor_boundary(floor):
    options = DB.Options()
    options.DetailLevel = DB.ViewDetailLevel.Fine
    geometry = floor.get_Geometry(options)
    for geometry_object in geometry:
        solid = geometry_object if isinstance(geometry_object, DB.Solid) else None
        if not solid or solid.Faces.Size == 0:
            continue
        for face in solid.Faces:
            planar = face if isinstance(face, DB.PlanarFace) else None
            if planar and planar.FaceNormal.Z > 0.9:
                loops = planar.GetEdgesAsCurveLoops()
                if loops and loops.Count > 0:
                    return loops[0]
    return None


def smart_safety_detailer():
    reference = revit.uidoc.Selection.PickObject(UI.Selection.ObjectType.Element, "Pick a floor or slab")
    floor = revit.doc.GetElement(reference.ElementId)
    if not floor or not isinstance(floor, DB.Floor):
        forms.alert("Please select a floor or slab element.", title="Aurora Smart Safety Detailer")
        return
    loop = _floor_boundary(floor)
    if not loop:
        forms.alert("Could not find a planar floor boundary for preview.", title="Aurora Smart Safety Detailer")
        return
    spacing_text = forms.ask_for_string(default="1.2", prompt="Post spacing in metres", title="Aurora Smart Safety Detailer")
    try:
        spacing = float(spacing_text or "1.2")
        if spacing <= 0:
            raise ValueError()
    except Exception:
        forms.alert("Spacing must be a positive number.", title="Aurora Smart Safety Detailer")
        return
    length = 0.0
    for curve in loop:
        try:
            length += curve.Length
        except Exception:
            pass
    preview = "Safe Preview\n\nFloor: {0}\nBoundary length: {1:.2f} ft\nRequested post spacing: {2:.2f} m\nApproximate posts: {3}\n\nThe loaded Revit railing type controls final baluster placement. Continue?".format(
        floor.Id.IntegerValue, length, spacing, int((length * 0.3048) / spacing) + 1)
    if not _confirm(preview, "Aurora Smart Safety Detailer"):
        return
    try:
        architecture = getattr(DB, "Architecture")
        railing_class = getattr(architecture, "Railing")
        railing_type_class = getattr(architecture, "RailingType")
        railing_type = DB.FilteredElementCollector(revit.doc).OfClass(railing_type_class).FirstElement()
    except Exception:
        railing_class = None
        railing_type = None
    if railing_class is None or railing_type is None:
        forms.alert("No supported railing type/API was available. Load a railing family/type and try again.", title="Aurora Smart Safety Detailer")
        return
    level_id = floor.LevelId
    tx = DB.Transaction(revit.doc, "Aurora Smart Safety Detailer")
    tx.Start()
    created = 0
    try:
        try:
            railing_class.Create(revit.doc, loop, railing_type.Id, level_id)
            created = 1
        except Exception:
            for curve in loop:
                single_loop = DB.CurveLoop()
                single_loop.Append(curve)
                railing_class.Create(revit.doc, single_loop, railing_type.Id, level_id)
                created += 1
        tx.Commit()
    except Exception:
        tx.RollBack()
        raise
    _log("SMART_SAFETY_DETAILER", "Created {0} railing path(s) for floor {1}; requested spacing {2} m".format(created, floor.Id.IntegerValue, spacing))
    forms.alert("Created {0} railing path(s). Verify the loaded railing type's baluster spacing in the model.".format(created), title="Aurora Smart Safety Detailer")


def main():
    if not forms or not revit or not DB or not UI:
        return
    choice = _choose("Aurora Utility Tools", [
        "Export Current View to PDF",
        "Export Active Schedule to Excel",
        "Batch Parameter Translator",
        "Background Calculation Killer",
        "Restore Calculations",
        "Smart Safety Detailer",
        "Schedule Export to Excel",
    ])
    try:
        if choice == "Export Current View to PDF":
            export_current_view_to_pdf()
        elif choice == "Export Active Schedule to Excel":
            export_active_schedule(True)
        elif choice == "Schedule Export to Excel":
            export_active_schedule(False)
        elif choice == "Batch Parameter Translator":
            batch_parameter_translator()
        elif choice == "Background Calculation Killer":
            background_calculation_killer()
        elif choice == "Restore Calculations":
            restore_calculations()
        elif choice == "Smart Safety Detailer":
            smart_safety_detailer()
    except Exception as error:
        forms.alert("Utility command failed safely:\n\n" + _text(error), title="Aurora Utility Tools")


if __name__ == "__main__":
    main()
